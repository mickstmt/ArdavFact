"""Rutas para la importación de comprobantes históricos (sin envío a SUNAT)."""
import io
import os
import uuid
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

from flask import render_template, request, jsonify, current_app, send_file
from flask_login import login_required, current_user

from app.decorators import requiere_permiso
from app.extensions import db
from app.models.comprobante import Comprobante, ComprobanteItem
from app.models.cliente import Cliente
from app.services.utils import calcular_igv_item
from . import historico_bp

_ALLOWED_EXT = {'xlsx', 'xls'}
_TIPOS_COMPROBANTE = {'BOLETA', 'FACTURA', 'NOTA_CREDITO'}
_TIPOS_DOC = {'DNI', 'RUC', 'CE', 'PAS', 'PASAPORTE'}
_TIPO_DOC_SUNAT = {'BOLETA': '03', 'FACTURA': '01', 'NOTA_CREDITO': '07'}
_D0 = Decimal('0')
_D2 = Decimal('0.01')


def _allowed(f: str) -> bool:
    return '.' in f and f.rsplit('.', 1)[1].lower() in _ALLOWED_EXT


def _serie_historica(serie: str) -> str:
    """B001→H001, F001→HF01, BC01→HC01, FC01→HFC1."""
    s = (serie or 'B001').strip().upper()
    return 'H' + s[1:]


def _norm_tipo_doc(raw: str) -> str:
    raw = (raw or '').strip().upper()
    if raw == 'PAS':
        return 'PASAPORTE'
    return raw if raw in _TIPOS_DOC else 'DNI'


def _parse_fecha(val, pandas) -> datetime | None:
    if pandas.isna(val) or str(val).strip() == '':
        return None
    if hasattr(val, 'to_pydatetime'):
        return val.to_pydatetime()
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Upload
# ─────────────────────────────────────────────────────────────────────────────

@historico_bp.route('/', methods=['GET'])
@login_required
@requiere_permiso('ventas.crear')
def upload():
    return render_template('historico/upload.html')


# ─────────────────────────────────────────────────────────────────────────────
# Analizar
# ─────────────────────────────────────────────────────────────────────────────

@historico_bp.route('/analizar', methods=['POST'])
@login_required
@requiere_permiso('ventas.crear')
def analizar():
    if 'archivo' not in request.files:
        return jsonify({'success': False, 'message': 'No se recibió archivo.'}), 400

    archivo = request.files['archivo']
    if not archivo.filename or not _allowed(archivo.filename):
        return jsonify({'success': False, 'message': 'Formato no válido. Use .xlsx o .xls.'}), 400

    uploads_path = current_app.config.get('UPLOADS_PATH', 'uploads')
    os.makedirs(uploads_path, exist_ok=True)
    tmp = os.path.join(uploads_path, f'hist_{uuid.uuid4().hex}.xlsx')

    try:
        archivo.save(tmp)
        ordenes = _analizar_excel(tmp)
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception as exc:
        current_app.logger.error('[HIST] Error analizando: %s', exc, exc_info=True)
        return jsonify({'success': False, 'message': f'Error al procesar el archivo: {exc}'}), 500
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)

    total_ok      = sum(1 for o in ordenes if o['status'] == 'OK')
    total_warning = sum(1 for o in ordenes if o['status'] == 'WARNING')
    total_error   = sum(1 for o in ordenes if o['status'] == 'ERROR')

    return jsonify({
        'success': True,
        'ordenes': ordenes,
        'resumen': {
            'total':   len(ordenes),
            'ok':      total_ok,
            'warning': total_warning,
            'error':   total_error,
        },
    })


def _analizar_excel(path: str) -> list:
    import pandas as pd

    df = pd.read_excel(path, dtype=str)

    # Normalizar cabeceras
    df.columns = [
        c.lower().strip()
         .replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ñ','n')
         .replace(' ', '_')
        for c in df.columns
    ]

    required = {'tipo_comprobante', 'serie', 'n_de_orden', 'fecha_emision',
                'tipo_doc', 'num_doc', 'nombre_cliente', 'nombre_producto',
                'cantidad', 'precio_unitario'}
    # Alias flexible para n de orden
    if 'n_de_orden' not in df.columns:
        for alt in ('n_orden', 'numero_orden', 'orden', 'numero_de_orden'):
            if alt in df.columns:
                df.rename(columns={alt: 'n_de_orden'}, inplace=True)
                break

    missing = required - set(df.columns)
    if missing:
        raise ValueError(f'Columnas faltantes: {", ".join(sorted(missing))}')

    grupos = {}  # numero_orden → orden dict

    for _, row in df.iterrows():
        def _v(col, default=''):
            val = row.get(col, default)
            if pd.isna(val):
                return default
            return str(val).strip()

        num_orden = _v('n_de_orden')
        if not num_orden:
            continue

        tipo_comp = _v('tipo_comprobante').upper()
        serie_orig = _v('serie', 'B001')
        fecha_raw  = row.get('fecha_emision')
        fecha      = _parse_fecha(fecha_raw, pd)
        tipo_doc   = _norm_tipo_doc(_v('tipo_doc'))
        num_doc    = _v('num_doc')
        nombre     = _v('nombre_cliente')
        costo_envio_raw = _v('costo_envio', '0')
        doc_ref    = _v('doc_referencia', '')
        total_raw  = _v('total_venta', '')

        if num_orden not in grupos:
            errores = []
            if tipo_comp not in _TIPOS_COMPROBANTE:
                errores.append(f'tipo_comprobante inválido: {tipo_comp!r}')
            if not fecha:
                errores.append('fecha_emision inválida o vacía')
            if not num_doc:
                errores.append('num_doc vacío')

            try:
                costo_envio = Decimal(costo_envio_raw.replace(',', '.') or '0')
            except Exception:
                costo_envio = _D0

            grupos[num_orden] = {
                'numero_orden':   num_orden,
                'tipo_comprobante': tipo_comp,
                'serie_original': serie_orig,
                'fecha_emision':  fecha.strftime('%Y-%m-%d') if fecha else '',
                'tipo_doc':       tipo_doc,
                'num_doc':        num_doc,
                'nombre_cliente': nombre,
                'costo_envio':    float(costo_envio),
                'doc_referencia': doc_ref,
                'total_manual':   total_raw,
                'items':          [],
                'errores':        errores,
                'advertencias':   [],
                'status':         'ERROR' if errores else 'OK',
            }

        orden = grupos[num_orden]

        # Ítem
        nom_prod = _v('nombre_producto')
        sku      = _v('sku', '')
        try:
            cantidad = Decimal(_v('cantidad', '1').replace(',', '.') or '1')
        except Exception:
            cantidad = Decimal('1')
        try:
            precio_u = Decimal(_v('precio_unitario', '0').replace(',', '.') or '0')
        except Exception:
            precio_u = _D0

        item_err = None
        if precio_u <= 0:
            item_err = f'Precio inválido ({precio_u}) para ítem "{nom_prod}"'
            if orden['status'] != 'ERROR':
                orden['status'] = 'ERROR'
            orden['errores'].append(item_err)
        if not nom_prod:
            item_err = 'nombre_producto vacío'
            if orden['status'] != 'ERROR':
                orden['status'] = 'ERROR'
            orden['errores'].append(item_err)

        igv = calcular_igv_item(precio_u, cantidad, '10')
        orden['items'].append({
            'nombre_producto':        nom_prod,
            'sku':                    sku,
            'cantidad':               float(cantidad),
            'precio_unitario_con_igv': float(precio_u),
            'precio_unitario_sin_igv': float(igv['precio_sin_igv']),
            'subtotal_con_igv':        float(igv['subtotal_con_igv']),
            'subtotal_sin_igv':        float(igv['subtotal_sin_igv']),
            'igv_unitario':            float(igv['igv_unitario']),
            'igv_total':               float(igv['igv_total']),
            'error':                   item_err,
        })

    # Calcular totales por orden
    ordenes = []
    for o in grupos.values():
        subtotal = sum(Decimal(str(i['subtotal_con_igv'])) for i in o['items'])
        costo    = Decimal(str(o['costo_envio']))
        if o['total_manual']:
            try:
                total = Decimal(o['total_manual'].replace(',', '.'))
            except Exception:
                total = subtotal + costo
        else:
            total = subtotal + costo
        o['subtotal'] = float(subtotal)
        o['total']    = float(total)
        o.pop('total_manual', None)
        ordenes.append(o)

    return ordenes


# ─────────────────────────────────────────────────────────────────────────────
# Preview
# ─────────────────────────────────────────────────────────────────────────────

@historico_bp.route('/preview', methods=['GET'])
@login_required
@requiere_permiso('ventas.crear')
def preview():
    return render_template('historico/preview.html')


# ─────────────────────────────────────────────────────────────────────────────
# Procesar
# ─────────────────────────────────────────────────────────────────────────────

@historico_bp.route('/procesar', methods=['POST'])
@login_required
@requiere_permiso('ventas.crear')
def procesar():
    payload = request.get_json(force=True) or {}
    ordenes = payload.get('ordenes', [])

    if not ordenes:
        return jsonify({'success': False, 'message': 'Sin órdenes para procesar.'}), 400

    try:
        creados = 0
        errores = 0

        # Pre-calcular siguientes correlativos por serie histórica
        series_counter = {}

        # Ordenar por fecha para asignar correlativos cronológicamente
        ordenes_sorted = sorted(
            ordenes,
            key=lambda o: o.get('fecha_emision', '') or ''
        )

        for o in ordenes_sorted:
            try:
                serie_hist = _serie_historica(o.get('serie_original', 'B001'))
                tipo_comp  = o['tipo_comprobante']

                # Correlativo siguiente para esta serie
                if serie_hist not in series_counter:
                    from sqlalchemy import func, cast, Integer
                    max_corr = db.session.query(func.max(cast(Comprobante.correlativo, Integer))) \
                        .filter(Comprobante.serie == serie_hist).scalar() or 0
                    series_counter[serie_hist] = max_corr

                series_counter[serie_hist] += 1
                corr     = str(series_counter[serie_hist]).zfill(8)
                num_comp = f'{serie_hist}-{corr}'

                # Cliente — upsert
                cliente = _upsert_cliente(
                    tipo_doc=o['tipo_doc'],
                    num_doc=o['num_doc'],
                    nombre=o['nombre_cliente'],
                )

                # Fecha
                fecha = datetime.strptime(o['fecha_emision'], '%Y-%m-%d') if o['fecha_emision'] else datetime.now()

                # Comprobante referencia (para NC)
                comp_ref_id = None
                if tipo_comp == 'NOTA_CREDITO' and o.get('doc_referencia'):
                    ref = Comprobante.query.filter_by(numero_completo=o['doc_referencia']).first()
                    if ref:
                        comp_ref_id = ref.id

                # Calcular totales tributarios
                items_dec = [_item_to_decimal(i) for i in o.get('items', [])]
                costo_envio = Decimal(str(o.get('costo_envio', 0) or 0))
                totales = _calcular_totales(items_dec, costo_envio)

                comp = Comprobante(
                    tipo_comprobante         = tipo_comp,
                    tipo_documento_sunat     = _TIPO_DOC_SUNAT.get(tipo_comp, '03'),
                    serie                    = serie_hist,
                    correlativo              = corr,
                    numero_completo          = num_comp,
                    numero_orden             = o.get('numero_orden'),
                    cliente_id               = cliente.id,
                    vendedor_id              = current_user.id,
                    subtotal                 = totales['subtotal'],
                    costo_envio              = costo_envio,
                    descuento                = _D0,
                    total_operaciones_gravadas = totales['total_gravadas'],
                    total_igv                = totales['total_igv'],
                    total                    = totales['total'],
                    estado                   = 'HISTORICO',
                    fecha_emision            = fecha,
                    fecha_pedido             = fecha,
                    es_bulk                  = False,
                    comprobante_referencia_id = comp_ref_id,
                )
                db.session.add(comp)
                db.session.flush()

                for item in o.get('items', []):
                    ci = ComprobanteItem(
                        comprobante_id           = comp.id,
                        producto_nombre          = item['nombre_producto'],
                        producto_sku             = item.get('sku') or None,
                        cantidad                 = Decimal(str(item['cantidad'])),
                        precio_unitario_con_igv  = Decimal(str(item['precio_unitario_con_igv'])),
                        precio_unitario_sin_igv  = Decimal(str(item['precio_unitario_sin_igv'])),
                        igv_unitario             = Decimal(str(item['igv_unitario'])),
                        subtotal_con_igv         = Decimal(str(item['subtotal_con_igv'])),
                        subtotal_sin_igv         = Decimal(str(item['subtotal_sin_igv'])),
                        igv_total                = Decimal(str(item['igv_total'])),
                        tipo_afectacion_igv      = '10',
                    )
                    db.session.add(ci)

                creados += 1

            except Exception as exc:
                current_app.logger.error('[HIST] Error orden %s: %s', o.get('numero_orden'), exc, exc_info=True)
                errores += 1

        db.session.commit()

    except Exception as exc:
        db.session.rollback()
        current_app.logger.error('[HIST] Error general: %s', exc, exc_info=True)
        return jsonify({'success': False, 'message': f'Error interno: {exc}'}), 500

    return jsonify({
        'success':  True,
        'creados':  creados,
        'errores':  errores,
        'message':  f'{creados} comprobante(s) creado(s)' + (f', {errores} con error' if errores else ''),
    })


def _upsert_cliente(tipo_doc: str, num_doc: str, nombre: str) -> Cliente:
    cliente = Cliente.query.filter_by(numero_documento=num_doc).first()
    if cliente:
        return cliente

    tipo_doc_af = tipo_doc if tipo_doc in ('DNI', 'RUC', 'CE', 'PASAPORTE') else 'DNI'

    cliente = Cliente(tipo_documento=tipo_doc_af, numero_documento=num_doc)

    if tipo_doc_af == 'RUC':
        cliente.razon_social = nombre
    else:
        partes = nombre.split()
        if len(partes) >= 3:
            cliente.nombres          = ' '.join(partes[:-2])
            cliente.apellido_paterno = partes[-2]
            cliente.apellido_materno = partes[-1]
        elif len(partes) == 2:
            cliente.nombres          = partes[0]
            cliente.apellido_paterno = partes[1]
        else:
            cliente.nombres = nombre

    db.session.add(cliente)
    db.session.flush()
    return cliente


class _ItemDec:
    """Objeto mínimo compatible con calcular_totales_comprobante."""
    def __init__(self, d: dict):
        self.subtotal_sin_igv   = Decimal(str(d['subtotal_sin_igv']))
        self.igv_total          = Decimal(str(d['igv_total']))
        self.tipo_afectacion_igv = '10'


def _item_to_decimal(d: dict) -> _ItemDec:
    return _ItemDec(d)


def _calcular_totales(items: list, costo_envio: Decimal) -> dict:
    total_gravadas = sum((i.subtotal_sin_igv for i in items), _D0)
    total_igv      = sum((i.igv_total for i in items), _D0)

    if costo_envio > _D0:
        from app.services.utils import IGV_DIVISOR
        envio_sin_igv = (costo_envio / IGV_DIVISOR).quantize(_D2, ROUND_HALF_UP)
        total_gravadas += envio_sin_igv
        total_igv      += (costo_envio - envio_sin_igv).quantize(_D2, ROUND_HALF_UP)

    subtotal = sum(
        (Decimal(str(i.subtotal_sin_igv)) + Decimal(str(i.igv_total)) for i in items), _D0
    )
    total = subtotal + costo_envio

    return {
        'subtotal':      subtotal.quantize(_D2),
        'total_gravadas': total_gravadas.quantize(_D2),
        'total_igv':     total_igv.quantize(_D2),
        'total':         total.quantize(_D2),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Plantilla
# ─────────────────────────────────────────────────────────────────────────────

@historico_bp.route('/plantilla', methods=['GET'])
@login_required
@requiere_permiso('ventas.crear')
def plantilla():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Histórico'

    OBL_FILL = PatternFill('solid', fgColor='FFD5F5DC')
    OPT_FILL = PatternFill('solid', fgColor='FFFFF9C4')
    EJM_FILL = PatternFill('solid', fgColor='FFF5F5F5')
    THIN     = Side(style='thin', color='CCCCCC')
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Banner fila 1
    ws.merge_cells('A1:N1')
    b = ws.cell(row=1, column=1,
                value='PLANTILLA IMPORTACIÓN HISTÓRICA — ArdavFact  |  Una fila por ítem. Mismo "n de orden" = misma venta.')
    b.font      = Font(bold=True, color='FFFFFFFF', size=10)
    b.fill      = PatternFill('solid', fgColor='FF1e3a5f')
    b.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Leyenda fila 2
    ws.merge_cells('A2:N2')
    l = ws.cell(row=2, column=1,
                value='Verde = Obligatorio  |  Amarillo = Opcional  |  Fila 4 es de ejemplo, bórrala antes de subir')
    l.font      = Font(italic=True, size=9, color='FF333333')
    l.fill      = PatternFill('solid', fgColor='FFE8F0FE')
    l.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    # Fila 3: cabeceras
    campos = [
        ('tipo_comprobante', True,  'BOLETA, FACTURA o NOTA_CREDITO'),
        ('serie',            True,  'Ej: B001, F001. Se convertirá a H001, HF01, etc.'),
        ('n de orden',       True,  'Identificador de la orden. Varias filas = misma venta'),
        ('fecha_emision',    True,  'DD/MM/YYYY o YYYY-MM-DD'),
        ('tipo_doc',         True,  'DNI, RUC, CE o PAS'),
        ('num_doc',          True,  'Número de documento'),
        ('nombre_cliente',   True,  'Nombre completo o razón social'),
        ('nombre_producto',  True,  'Descripción del ítem'),
        ('cantidad',         True,  'Entero o decimal'),
        ('precio_unitario',  True,  'Precio unitario con IGV'),
        ('sku',              False, 'Código interno (puede ir vacío)'),
        ('costo_envio',      False, 'Costo de envío (0 si no aplica)'),
        ('doc_referencia',   False, 'Solo para NOTA_CREDITO. Ej: B001-00000001'),
        ('total_venta',      False, 'Total de la venta. Se calcula si se deja vacío'),
    ]
    for col_idx, (nombre, oblig, nota) in enumerate(campos, start=1):
        c = ws.cell(row=3, column=col_idx, value=nombre)
        c.font      = Font(bold=True, size=10)
        c.fill      = OBL_FILL if oblig else OPT_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = BORDER
        cm = Comment(nota, 'ArdavFact')
        cm.width, cm.height = 220, 50
        c.comment = cm
    ws.row_dimensions[3].height = 28

    # Filas 4-5: ejemplo
    ejemplos = [
        ['BOLETA', 'B001', 'ORD-001', '15/01/2026', 'DNI', '12345678', 'Juan Pérez',
         'Zapatilla Running Roja T42', '1', '118.00', '1234567-ROJO-42', '10.00', '', '128.00'],
        ['BOLETA', 'B001', 'ORD-001', '15/01/2026', 'DNI', '12345678', 'Juan Pérez',
         'Polo Básico Blanco T M', '2', '35.40', '9876543-BL-M', '0', '', ''],
    ]
    for row_off, fila in enumerate(ejemplos):
        for col_idx, val in enumerate(fila, start=1):
            c = ws.cell(row=4 + row_off, column=col_idx, value=val)
            c.font      = Font(italic=True, color='FF555555', size=9)
            c.fill      = EJM_FILL
            c.alignment = Alignment(vertical='center')
            c.border    = BORDER
        ws.row_dimensions[4 + row_off].height = 17

    anchos = [18, 8, 14, 14, 8, 14, 26, 32, 10, 16, 20, 12, 20, 14]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    # Hoja Leyenda
    ws2 = wb.create_sheet('Leyenda')
    leyenda = [
        ('tipo_comprobante', 'BOLETA, FACTURA o NOTA_CREDITO',                                     True),
        ('serie',            'Serie original. Ej: B001, F001. Se prefijará con H automáticamente.', True),
        ('n de orden',       'Identificador de la orden. Varias filas con el mismo valor = una venta.', True),
        ('fecha_emision',    'Fecha del comprobante. Formatos: DD/MM/YYYY, YYYY-MM-DD',              True),
        ('tipo_doc',         'DNI, RUC, CE (Carnet de Extranjería) o PAS (Pasaporte)',               True),
        ('num_doc',          'Número de documento del cliente',                                       True),
        ('nombre_cliente',   'Nombre completo o razón social',                                        True),
        ('nombre_producto',  'Descripción del ítem',                                                  True),
        ('cantidad',         'Cantidad vendida. Se acepta decimal',                                   True),
        ('precio_unitario',  'Precio unitario incluyendo IGV (18%)',                                  True),
        ('sku',              'SKU interno. Puede dejarse vacío',                                      False),
        ('costo_envio',      'Costo de envío de la orden. Solo en la primera fila de cada orden',     False),
        ('doc_referencia',   'Número del comprobante anulado. Solo para NOTA_CREDITO. Ej: B001-00000001', False),
        ('total_venta',      'Total de la venta. Si vacío, se calcula desde precio_unitario × cantidad', False),
    ]
    ws2.cell(row=1, column=1, value='Campo').font        = Font(bold=True)
    ws2.cell(row=1, column=2, value='Descripción').font  = Font(bold=True)
    ws2.cell(row=1, column=3, value='Obligatorio').font  = Font(bold=True)
    for i, (campo, desc, oblig) in enumerate(leyenda, start=2):
        ws2.cell(row=i, column=1, value=campo).fill = OBL_FILL if oblig else OPT_FILL
        ws2.cell(row=i, column=2, value=desc)
        ws2.cell(row=i, column=3, value='Sí' if oblig else 'No')
        for col in range(1, 4):
            ws2.cell(row=i, column=col).border    = Border(left=Side(style='thin', color='CCCCCC'),
                                                            right=Side(style='thin', color='CCCCCC'),
                                                            top=Side(style='thin', color='CCCCCC'),
                                                            bottom=Side(style='thin', color='CCCCCC'))
            ws2.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
        ws2.row_dimensions[i].height = 20
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 72
    ws2.column_dimensions['C'].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name='plantilla_historico_ardavfact.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
