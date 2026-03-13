"""Rutas de carga masiva desde Excel (WooCommerce, Falabella, MercadoLibre)."""
import io
import os
import uuid
import base64
from datetime import datetime

from flask import (
    render_template, request, jsonify, current_app,
    redirect, url_for, flash,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app.decorators import requiere_permiso
from app.services import bulk_service, bulk_falabella_service, bulk_meli_service
from . import bulk_bp

_SERVICIOS = {
    'woo':       bulk_service,
    'falabella': bulk_falabella_service,
    'meli':      bulk_meli_service,
}

_ALLOWED_EXT = {'xlsx', 'xls'}


def _allowed(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in _ALLOWED_EXT


# ─────────────────────────────────────────────────────────────────────────────
# Upload
# ─────────────────────────────────────────────────────────────────────────────

@bulk_bp.route('/', methods=['GET'])
@login_required
@requiere_permiso('ventas.crear')
def upload():
    """Formulario de carga masiva."""
    return render_template('bulk/upload.html')


@bulk_bp.route('/analizar', methods=['POST'])
@login_required
@requiere_permiso('ventas.crear')
def analizar():
    """Recibe el Excel, lo analiza y devuelve el preview en JSON."""
    if 'archivo' not in request.files:
        return jsonify({'success': False, 'message': 'No se recibió archivo.'}), 400

    archivo = request.files['archivo']
    if not archivo.filename or not _allowed(archivo.filename):
        return jsonify({'success': False, 'message': 'Formato no válido. Use .xlsx o .xls.'}), 400

    # Guardar temporalmente
    uploads_path = current_app.config.get('UPLOADS_PATH', 'uploads')
    os.makedirs(uploads_path, exist_ok=True)
    nombre_temp = f'bulk_{uuid.uuid4().hex}.xlsx'
    ruta_temp = os.path.join(uploads_path, nombre_temp)

    plataforma = request.form.get('plataforma', 'woo').strip().lower()
    servicio = _SERVICIOS.get(plataforma, bulk_service)

    try:
        archivo.save(ruta_temp)
        ordenes = servicio.analizar_excel(ruta_temp, current_app.config)
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception as exc:
        current_app.logger.error('[BULK] Error analizando Excel: %s', exc, exc_info=True)
        return jsonify({'success': False, 'message': f'Error al procesar el archivo: {exc}'}), 500
    finally:
        if os.path.exists(ruta_temp):
            os.remove(ruta_temp)

    total_ok      = sum(1 for o in ordenes if o['status'] == 'OK')
    total_warning = sum(1 for o in ordenes if o['status'] == 'WARNING')
    total_error   = sum(1 for o in ordenes if o['status'] == 'ERROR')

    return jsonify({
        'success':    True,
        'plataforma': plataforma,
        'ordenes':    ordenes,
        'resumen': {
            'total':   len(ordenes),
            'ok':      total_ok,
            'warning': total_warning,
            'error':   total_error,
        },
    })


# ─────────────────────────────────────────────────────────────────────────────
# Preview (página HTML)
# ─────────────────────────────────────────────────────────────────────────────

@bulk_bp.route('/preview', methods=['GET'])
@login_required
@requiere_permiso('ventas.crear')
def preview():
    """Página de preview (la tabla se carga vía JS con datos del localStorage)."""
    return render_template('bulk/preview.html')


# ─────────────────────────────────────────────────────────────────────────────
# Procesar
# ─────────────────────────────────────────────────────────────────────────────

@bulk_bp.route('/procesar', methods=['POST'])
@login_required
@requiere_permiso('ventas.crear')
def procesar():
    """Crea comprobantes para las órdenes seleccionadas (OK y WARNING)."""
    payload = request.get_json(force=True) or {}
    ordenes = payload.get('ordenes', [])
    fecha_override = (payload.get('fecha_override') or '').strip() or None

    if not ordenes:
        return jsonify({'success': False, 'message': 'Sin órdenes para procesar.'}), 400

    try:
        resultados = bulk_service.procesar_ordenes(
            ordenes,
            current_app.config,
            vendedor_id=current_user.id,
            fecha_override=fecha_override,
        )
    except Exception as exc:
        current_app.logger.error('[BULK] Error en procesar: %s', exc, exc_info=True)
        return jsonify({'success': False, 'message': f'Error interno: {exc}'}), 500

    exitosos = sum(1 for r in resultados if r.get('success'))
    fallidos  = len(resultados) - exitosos

    return jsonify({
        'success': True,
        'message': f'{exitosos} comprobante(s) creado(s), {fallidos} error(es).',
        'resultados': resultados,
        'exitosos': exitosos,
        'fallidos': fallidos,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Descargar plantilla
# ─────────────────────────────────────────────────────────────────────────────

@bulk_bp.route('/plantilla/<fuente>', methods=['GET'])
@login_required
@requiere_permiso('ventas.crear')
def descargar_plantilla(fuente):
    """Genera y descarga un Excel de plantilla con el formato exacto de cada fuente."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if fuente not in ('woo', 'falabella', 'meli'):
        return 'Fuente no válida.', 400

    wb   = openpyxl.Workbook()
    ws   = wb.active

    OBL_FILL  = PatternFill('solid', fgColor='FFD5F5DC')   # verde suave — obligatorio
    OPT_FILL  = PatternFill('solid', fgColor='FFFFF9C4')   # amarillo suave — opcional
    EJM_FILL  = PatternFill('solid', fgColor='FFF5F5F5')   # gris muy claro — ejemplo
    HDR_FONT  = Font(bold=True, size=10)
    EJM_FONT  = Font(italic=True, color='FF555555', size=9)
    THIN      = Side(style='thin', color='CCCCCC')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _hdr(ws, row, col, texto, oblig):
        c = ws.cell(row=row, column=col, value=texto)
        c.font      = HDR_FONT
        c.fill      = OBL_FILL if oblig else OPT_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = BORDER
        ws.row_dimensions[row].height = 30
        return c

    def _ejm(ws, row, col, valor):
        c = ws.cell(row=row, column=col, value=valor)
        c.font      = EJM_FONT
        c.fill      = EJM_FILL
        c.alignment = Alignment(vertical='center')
        c.border    = BORDER
        ws.row_dimensions[row].height = 18
        return c

    # ── Hoja de leyenda ───────────────────────────────────────────────────────
    ws_ley = wb.create_sheet('Leyenda')
    leyenda = [
        ('Color',        'Significado'),
        ('Verde',        'Campo OBLIGATORIO — debe estar completo en cada fila'),
        ('Amarillo',     'Campo OPCIONAL — puede dejarse vacío'),
        ('Gris itálica', 'Fila de EJEMPLO — borra esta fila antes de cargar'),
    ]
    fills_ley = [PatternFill('solid', fgColor='FFAAAAAA'), OBL_FILL, OPT_FILL, EJM_FILL]
    for i, (col, sig) in enumerate(leyenda, start=1):
        ca = ws_ley.cell(row=i, column=1, value=col)
        cb = ws_ley.cell(row=i, column=2, value=sig)
        for c in (ca, cb):
            c.fill   = fills_ley[i - 1]
            c.border = BORDER
            c.alignment = Alignment(vertical='center')
        ws_ley.row_dimensions[i].height = 18
    ws_ley.column_dimensions['A'].width = 16
    ws_ley.column_dimensions['B'].width = 58

    # ═════════════════════════════════════════════════════════════════════════
    # WooCommerce — header=3 → las 3 primeras filas son libres, datos desde f4
    # ═════════════════════════════════════════════════════════════════════════
    if fuente == 'woo':
        ws.title = 'WooCommerce'

        # Filas 1-3: banner informativo (pandas las ignora con header=3)
        ws.merge_cells('A1:X1')
        banner = ws.cell(row=1, column=1,
                         value='PLANTILLA CARGA MASIVA — WooCommerce  |  Cabeceras en fila 4, datos desde fila 5  |  Agrupar ítems por PedidoId (col. A)')
        banner.font      = Font(bold=True, color='FFFFFFFF', size=10)
        banner.fill      = PatternFill('solid', fgColor='FF1e3a5f')
        banner.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22

        ws.merge_cells('A2:X2')
        ley_cell = ws.cell(row=2, column=1,
                           value='Verde = Obligatorio  |  Amarillo = Opcional  |  La fila 5 es de ejemplo, bórrala antes de subir')
        ley_cell.font      = Font(italic=True, size=9, color='FF333333')
        ley_cell.alignment = Alignment(horizontal='center', vertical='center')
        ley_cell.fill      = PatternFill('solid', fgColor='FFE8F0FE')
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 8  # separador

        # Fila 4: cabeceras en posición exacta (1-indexed: A=1, C=3, G=7, H=8, I=9, J=10, K=11, M=13, X=24)
        campos = [
            (1,  'PedidoId\n(N° Orden)',   True),   # A
            (3,  'Fecha',                  True),   # C
            (7,  'DNI / RUC',              False),  # G
            (8,  'Nombre cliente',         False),  # H
            (9,  'Descripción',            False),  # I
            (10, 'SKU',                    True),   # J
            (11, 'Cantidad',               False),  # K  (default 1)
            (13, 'Precio ítem\n(con IGV)', True),   # M
            (24, 'Costo Envío\n(con IGV)', False),  # X
        ]
        for col, texto, oblig in campos:
            _hdr(ws, 4, col, texto, oblig)

        # Fila 5: ejemplo
        ejm = [
            (1,  '10001'),
            (3,  '15/01/2026'),
            (7,  '12345678'),
            (8,  'Juan Pérez'),
            (9,  'Zapatilla Running Roja T42'),
            (10, '1234567-ROJO-42'),
            (11, '1'),
            (13, '118.00'),
            (24, '10.00'),
        ]
        for col, val in ejm:
            _ejm(ws, 5, col, val)

        # Anchos de columnas relevantes
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 26
        ws.column_dimensions['I'].width = 34
        ws.column_dimensions['J'].width = 22
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['M'].width = 16
        ws.column_dimensions['X'].width = 16

        filename = 'plantilla_woocommerce.xlsx'

    # ═════════════════════════════════════════════════════════════════════════
    # Falabella — header=0 → fila 1 son cabeceras, datos desde fila 2
    # ═════════════════════════════════════════════════════════════════════════
    elif fuente == 'falabella':
        ws.title = 'Falabella'

        # Fila 1: cabeceras (B=2, D=4, E=5, J=10, L=12, AJ=36, AL=38, AO=41)
        campos = [
            (2,  'SKU',                    True),   # B
            (4,  'Fecha',                  True),   # D
            (5,  'N° Orden',               True),   # E
            (10, 'Nombre cliente',         False),  # J
            (12, 'DNI / RUC',              False),  # L
            (36, 'Precio ítem\n(con IGV)', True),   # AJ
            (38, 'Costo Envío\n(con IGV)', False),  # AL
            (41, 'Descripción',            False),  # AO
        ]
        for col, texto, oblig in campos:
            _hdr(ws, 1, col, texto, oblig)

        # Fila 2: ejemplo
        ejm = [
            (2,  '1234567-AZUL-38'),
            (4,  '15/01/2026'),
            (5,  '3001234567890'),
            (10, 'Ana García'),
            (12, '45678901'),
            (36, '89.90'),
            (38, '5.00'),
            (41, 'Polo Manga Corta Azul T S'),
        ]
        for col, val in ejm:
            _ejm(ws, 2, col, val)

        for col_letter, w in [('B',22),('D',14),('E',18),('J',26),('L',14),
                               (get_column_letter(36),16),(get_column_letter(38),16),(get_column_letter(41),34)]:
            ws.column_dimensions[col_letter].width = w

        filename = 'plantilla_falabella.xlsx'

    # ═════════════════════════════════════════════════════════════════════════
    # MercadoLibre — header=0, datos generales solo en primera fila de cada orden
    # ═════════════════════════════════════════════════════════════════════════
    else:  # meli
        ws.title = 'MercadoLibre'

        # A=1, G=7, V=22, AE=31, AF=32, AH=34, AI=35, AQ=43, AU=47, AX=50
        campos = [
            (1,  'N° Orden\n(1ª fila del grupo)', True),   # A
            (7,  'Fecha\n(1ª fila del grupo)',     True),   # G
            (22, 'Costo Envío\n(1ª fila, c/IGV)',  False),  # V
            (31, 'SKU',                            True),   # AE
            (32, 'Variante\n(color/talla)',         False),  # AF
            (34, 'Precio ítem\n(con IGV)',          True),   # AH
            (35, 'Cantidad',                        False),  # AI
            (43, 'Descripción',                     False),  # AQ
            (47, 'Nombre cliente\n(1ª fila)',        False),  # AU
            (50, 'DNI / RUC\n(1ª fila)',             False),  # AX
        ]
        for col, texto, oblig in campos:
            _hdr(ws, 1, col, texto, oblig)

        # Fila 2 (primera fila de una orden — lleva datos generales + primer ítem)
        ejm1 = [
            (1,  '2000011889490605'),
            (7,  '15/01/2026'),
            (22, '8.00'),
            (31, '9876543-NEGRO-M'),
            (32, 'Negro / M'),
            (34, '75.50'),
            (35, '2'),
            (43, 'Casaca Deportiva Negro Talla M'),
            (47, 'Carlos López'),
            (50, '71234567'),
        ]
        for col, val in ejm1:
            _ejm(ws, 2, col, val)

        # Fila 3 (segunda fila de la misma orden — solo SKU y datos del ítem)
        ejm2 = [
            (1,  '2000011889490605'),  # mismo N° Orden
            (31, '1122334-BLANCO-L'),
            (34, '65.00'),
            (35, '1'),
            (43, 'Polo Básico Blanco Talla L'),
        ]
        for col, val in ejm2:
            _ejm(ws, 3, col, val)

        for col_n, w in [(1,22),(7,14),(22,16),(31,22),(32,14),(34,16),(35,10),(43,34),(47,26),(50,14)]:
            ws.column_dimensions[get_column_letter(col_n)].width = w

        # Nota aclaratoria en la fila de ejemplo
        nota = ws.cell(row=2, column=53,
                       value='← Datos del cliente y envío SOLO en la 1ª fila de cada orden')
        nota.font      = Font(italic=True, size=8, color='FF888888')
        nota.alignment = Alignment(vertical='center')

        filename = 'plantilla_mercadolibre.xlsx'

    ws.freeze_panes = ws.cell(row=2 if fuente != 'woo' else 5, column=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ─────────────────────────────────────────────────────────────────────────────
# Descargar errores
# ─────────────────────────────────────────────────────────────────────────────

@bulk_bp.route('/descargar-errores', methods=['POST'])
@login_required
@requiere_permiso('ventas.crear')
def descargar_errores():
    """Genera Excel de errores en el mismo layout que el archivo fuente original."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    payload = request.get_json(force=True) or {}
    ordenes = payload.get('ordenes', [])
    fuente  = payload.get('fuente', 'woo')

    if not ordenes:
        return jsonify({'success': False, 'message': 'Sin datos para exportar.'}), 400

    # ── Layouts por fuente (columnas 1-based) ─────────────────────────────────
    # Cada layout define exactamente en qué columna va cada dato,
    # igual que el archivo original — listo para corregir y re-subir.
    #
    # Falabella / MercadoLibre comparten cabecera en fila 1, datos desde fila 2
    # WooCommerce: cabecera en fila 4, datos desde fila 5

    LAYOUTS = {
        'falabella': {
            'header_row': 1, 'data_start': 2,
            'freeze': 'B2',
            'col_orden':   5,   # E
            'col_fecha':   4,   # D
            'col_nombre':  10,  # J
            'col_doc':     12,  # L
            'col_sku':     2,   # B
            'col_desc':    41,  # AO
            'col_precio':  36,  # AJ
            'col_envio':   38,  # AL
            'col_cant':    None,
            'col_error':   42,  # AP
            'col_widths':  {2:18, 5:22, 10:28, 12:16, 36:16, 38:16, 41:34, 42:50},
        },
        'meli': {
            'header_row': 1, 'data_start': 2,
            'freeze': 'B2',
            'col_orden':   1,   # A
            'col_fecha':   7,   # G
            'col_nombre':  47,  # AU
            'col_doc':     50,  # AX
            'col_sku':     31,  # AE
            'col_desc':    43,  # AQ
            'col_precio':  34,  # AH
            'col_envio':   22,  # V
            'col_cant':    35,  # AI
            'col_error':   51,  # AY
            'col_widths':  {1:22, 7:14, 22:16, 31:22, 34:16, 35:10, 43:34, 47:26, 50:14, 51:50},
        },
        'woo': {
            'header_row': 4, 'data_start': 5,
            'freeze': 'A5',
            'col_orden':   1,   # A
            'col_fecha':   3,   # C
            'col_nombre':  8,   # H
            'col_doc':     7,   # G
            'col_sku':     10,  # J
            'col_desc':    9,   # I
            'col_precio':  13,  # M
            'col_envio':   24,  # X
            'col_cant':    11,  # K
            'col_error':   25,  # Y
            'col_widths':  {1:22, 3:14, 7:16, 8:28, 9:34, 10:22, 11:10, 13:16, 24:16, 25:50},
        },
    }

    layout = LAYOUTS.get(fuente, LAYOUTS['woo'])
    HDR_ROW = layout['header_row']

    HDR_FILL  = PatternFill('solid', fgColor='FF1e3a5f')
    ERR_FILL  = PatternFill('solid', fgColor='FFF8D7DA')
    WARN_FILL = PatternFill('solid', fgColor='FFFFF3CD')
    THIN      = Side(style='thin', color='CCCCCC')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = {'woo': 'WooCommerce', 'meli': 'MercadoLibre', 'falabella': 'Falabella'}.get(fuente, 'Errores')

    # ── Cabeceras ─────────────────────────────────────────────────────────────
    col_labels = {
        layout['col_orden']:  'N° Orden',
        layout['col_fecha']:  'Fecha',
        layout['col_nombre']: 'Nombre Cliente',
        layout['col_doc']:    'DNI / RUC',
        layout['col_sku']:    'SKU',
        layout['col_desc']:   'Descripción',
        layout['col_precio']: 'Precio (S/)',
        layout['col_envio']:  'Costo Envío (S/)',
        layout['col_error']:  'ERROR / ADVERTENCIA  ← corregir aquí',
    }
    if layout['col_cant']:
        col_labels[layout['col_cant']] = 'Cantidad'

    for col, label in col_labels.items():
        cell = ws.cell(row=HDR_ROW, column=col, value=label)
        cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[HDR_ROW].height = 28

    for col, w in layout['col_widths'].items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Datos ─────────────────────────────────────────────────────────────────
    def _celda(row, col, valor, fill=None):
        if col is None:
            return
        c = ws.cell(row=row, column=col, value=valor)
        if fill:
            c.fill = fill
        c.border    = BORDER
        c.alignment = Alignment(vertical='center', wrap_text=(col == layout['col_error']))
        return c

    current_row = layout['data_start']

    for orden in ordenes:
        num_orden    = orden.get('numero_orden', '')
        fecha_str    = (orden.get('fecha_emision') or '')[:10]
        nombre       = orden.get('nombre_cliente', '')
        num_doc      = orden.get('numero_documento', '')
        costo_envio  = float(orden.get('costo_envio', 0) or 0)
        errores_ord  = orden.get('errores', [])
        advertencias = orden.get('advertencias', [])
        items        = orden.get('items', [])
        status       = orden.get('status', 'ERROR')
        ord_fill     = ERR_FILL if status == 'ERROR' else WARN_FILL

        if not items:
            detalle = ' | '.join(errores_ord + advertencias) or '—'
            _celda(current_row, layout['col_orden'],  num_orden, ord_fill)
            _celda(current_row, layout['col_fecha'],  fecha_str, ord_fill)
            _celda(current_row, layout['col_nombre'], nombre,    ord_fill)
            _celda(current_row, layout['col_doc'],    num_doc,   ord_fill)
            _celda(current_row, layout['col_envio'],  costo_envio, ord_fill)
            _celda(current_row, layout['col_error'],  detalle,   ord_fill)
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            continue

        primera_fila = True
        for item in items:
            item_error = item.get('error')
            item_adv   = item.get('advertencia')
            if item_error:
                detalle = item_error
                fill    = ERR_FILL
            elif item_adv:
                detalle = item_adv
                fill    = WARN_FILL
            elif errores_ord or advertencias:
                detalle = ' | '.join(errores_ord + advertencias)
                fill    = ord_fill
            else:
                continue

            _celda(current_row, layout['col_sku'],    item.get('sku', ''),                    fill)
            _celda(current_row, layout['col_desc'],   item.get('descripcion', ''),            fill)
            _celda(current_row, layout['col_precio'],  float(item.get('precio_con_igv', 0) or 0), fill)
            _celda(current_row, layout['col_cant'],   float(item.get('cantidad', 1) or 1),    fill)
            _celda(current_row, layout['col_error'],  detalle,                                fill)
            # Datos de orden solo en primera fila del grupo
            if primera_fila:
                _celda(current_row, layout['col_orden'],  num_orden,   fill)
                _celda(current_row, layout['col_fecha'],  fecha_str,   fill)
                _celda(current_row, layout['col_nombre'], nombre,      fill)
                _celda(current_row, layout['col_doc'],    num_doc,     fill)
                _celda(current_row, layout['col_envio'],  costo_envio, fill)
                primera_fila = False

            ws.row_dimensions[current_row].height = 18
            current_row += 1

    ws.freeze_panes = layout['freeze']

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    b64 = base64.b64encode(output.read()).decode('utf-8')
    nombre_archivo = f'errores_{fuente}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return jsonify({'success': True, 'filename': nombre_archivo, 'filedata': b64})
