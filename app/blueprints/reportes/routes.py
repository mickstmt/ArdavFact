"""Rutas de reportes: ganancias con desglose IGV y exportación Excel."""
import io
from datetime import datetime, date
from decimal import Decimal

from flask import render_template, request, send_file, current_app
from flask_login import login_required
from sqlalchemy import func

from app.extensions import db
from app.models.comprobante import Comprobante
from app.models.producto import CostoProducto
from app.decorators import requiere_permiso
from app.services.utils import extraer_skus_base
from app.services.tipo_cambio_service import get_tipo_cambio
from . import reportes_bp


def _build_mapa_costos() -> dict:
    """Carga todos los costos en memoria: {sku_str: costo_usd}."""
    mapa = {}
    for c in CostoProducto.query.all():
        sku_str = str(c.sku).split('.')[0].strip()
        if sku_str:
            mapa[sku_str] = float(c.costo)
    return mapa

_TIPOS_VENTA     = ('FACTURA', 'BOLETA')
_ESTADOS_VALIDOS = ('ENVIADO', 'ACEPTADO', 'HISTORICO')

# Patrones de numero_orden por fuente
_FUENTE_FILTROS = {
    'MercadoLibre': lambda q: q.filter(
        Comprobante.numero_orden.ilike('2%'),
        func.length(Comprobante.numero_orden) >= 14,
    ),
    'Falabella': lambda q: q.filter(
        Comprobante.numero_orden.ilike('3%'),
        func.length(Comprobante.numero_orden) >= 10,
    ),
    'WooCommerce': lambda q: q.filter(
        Comprobante.numero_orden.isnot(None),
        ~Comprobante.numero_orden.ilike('2%') | (func.length(Comprobante.numero_orden) < 14),
        ~Comprobante.numero_orden.ilike('3%') | (func.length(Comprobante.numero_orden) < 10),
    ),
    'Manual': lambda q: q.filter(Comprobante.numero_orden.is_(None)),
}


# ─────────────────────────────────────────────────────────────────────────────
# Reporte de Ganancias
# ─────────────────────────────────────────────────────────────────────────────

@reportes_bp.route('/ganancias')
@login_required
@requiere_permiso('reportes.ver')
def ganancias():
    """Dashboard financiero con desglose IGV y tabla por comprobante."""
    fecha_ini_str = request.args.get('fecha_ini', '')
    fecha_fin_str = request.args.get('fecha_fin', '')
    tipo_filtro   = request.args.get('tipo', '')
    page          = request.args.get('page', 1, type=int)

    fuente_filtro = request.args.get('fuente', '')

    fecha_ini = _parse_date(fecha_ini_str)
    fecha_fin = _parse_date(fecha_fin_str)

    hoy = date.today()
    if not fecha_ini:
        fecha_ini = hoy.replace(day=1)
    if not fecha_fin:
        fecha_fin = hoy

    query = (
        Comprobante.query
        .filter(
            Comprobante.tipo_comprobante.in_(_TIPOS_VENTA),
            Comprobante.estado.in_(_ESTADOS_VALIDOS),
            func.date(Comprobante.fecha_emision) >= fecha_ini,
            func.date(Comprobante.fecha_emision) <= fecha_fin,
        )
    )
    if tipo_filtro in _TIPOS_VENTA:
        query = query.filter(Comprobante.tipo_comprobante == tipo_filtro)
    if fuente_filtro in _FUENTE_FILTROS:
        query = _FUENTE_FILTROS[fuente_filtro](query)

    mapa_costos = _build_mapa_costos()
    todos     = query.all()
    resumen   = _calcular_resumen(todos, mapa_costos)
    paginated = query.order_by(Comprobante.fecha_emision.desc()).paginate(
        page=page, per_page=30, error_out=False
    )
    filas = [_enriquecer_fila(c, mapa_costos) for c in paginated.items]

    return render_template(
        'reportes/ganancias.html',
        resumen=resumen,
        filas=filas,
        comprobantes=paginated,
        fecha_ini=fecha_ini.isoformat(),
        fecha_fin=fecha_fin.isoformat(),
        tipo_filtro=tipo_filtro,
        fuente_filtro=fuente_filtro,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Exportación Excel
# ─────────────────────────────────────────────────────────────────────────────

@reportes_bp.route('/ganancias/exportar')
@login_required
@requiere_permiso('reportes.exportar')
def exportar_ganancias():
    """Exporta reporte a Excel (.xlsx) con hoja de detalle + resumen fiscal."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    fecha_ini     = _parse_date(request.args.get('fecha_ini', '')) or date.today().replace(day=1)
    fecha_fin     = _parse_date(request.args.get('fecha_fin', '')) or date.today()
    tipo_filtro   = request.args.get('tipo', '')
    fuente_filtro = request.args.get('fuente', '')

    query = (
        Comprobante.query
        .filter(
            Comprobante.tipo_comprobante.in_(_TIPOS_VENTA),
            Comprobante.estado.in_(_ESTADOS_VALIDOS),
            func.date(Comprobante.fecha_emision) >= fecha_ini,
            func.date(Comprobante.fecha_emision) <= fecha_fin,
        )
        .order_by(Comprobante.fecha_emision)
    )
    if tipo_filtro in _TIPOS_VENTA:
        query = query.filter(Comprobante.tipo_comprobante == tipo_filtro)
    if fuente_filtro in _FUENTE_FILTROS:
        query = _FUENTE_FILTROS[fuente_filtro](query)

    mapa_costos = _build_mapa_costos()
    todos   = query.all()
    resumen = _calcular_resumen(todos, mapa_costos)
    filas   = [_enriquecer_fila(c, mapa_costos) for c in todos]

    wb = openpyxl.Workbook()

    # ── Hoja 1: Detalle ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Detalle'
    _COLOR_HDR = 'FF1e3a5f'

    # cols: 1=Fuente 2=N°Orden 3=Serie 4=Correlativo 5=Tipo 6=Fecha 7=Cliente 8=Estado
    #       9=Base 10=IGV 11=Descuento 12=CostoEnvío 13=Total
    #       14=CostoProductos 15=GananciaBruta 16=Margen
    hdrs = [
        'Fuente', 'N° Orden', 'Serie', 'Correlativo', 'Tipo', 'Fecha', 'Cliente', 'Estado',
        'Base Imponible', 'IGV 18%', 'Descuento', 'Costo Envío', 'Total Ingreso',
        'Costo Productos', 'Ganancia Bruta', 'Margen %',
    ]
    ws.append(hdrs)
    for col_idx in range(1, len(hdrs) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color='FFFFFFFF')
        cell.fill      = PatternFill('solid', fgColor=_COLOR_HDR)
        cell.alignment = Alignment(horizontal='center')

    for f in filas:
        partes = f['numero_completo'].split('-', 1)
        ws.append([
            f['fuente'],
            f['numero_orden'],
            partes[0] if len(partes) > 0 else f['numero_completo'],
            partes[1] if len(partes) > 1 else '',
            f['tipo_comprobante'],
            f['fecha_emision'],
            f['cliente_nombre'],
            f['estado'],
            float(f['base_imponible']),
            float(f['total_igv']),
            float(f['descuento']),
            float(f['costo_envio']),
            float(f['total']),
            float(f['costo_productos']),
            float(f['ganancia_bruta']),
            float(f['margen_pct']),
        ])

    moneda_cols = {9, 10, 11, 12, 13, 14, 15}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in moneda_cols:
                cell.number_format = '"S/ "#,##0.00'
            elif cell.column == 16:
                cell.number_format = '0.00"%"'

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max((len(str(c.value)) for c in col if c.value), default=10) + 4, 40
        )

    # ── Hoja 2: Resumen Fiscal ───────────────────────────────────────────────
    ws2 = wb.create_sheet('Resumen Fiscal')
    cfg  = current_app.config
    ws2.append(['RESUMEN FISCAL'])
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.append([f'{cfg.get("EMPRESA_RAZON_SOCIAL","")} — RUC {cfg.get("EMPRESA_RUC","")}'])
    ws2.append([f'Período: {fecha_ini.strftime("%d/%m/%Y")} — {fecha_fin.strftime("%d/%m/%Y")}'])
    ws2.append([])
    ws2.append(['Concepto', 'Monto (S/)'])
    ws2['A5'].font = ws2['B5'].font = Font(bold=True)

    for concepto, monto in [
        ('Total Ingresos (con IGV)',  resumen['total_ingresos']),
        ('Base Imponible (sin IGV)',  resumen['base_imponible']),
        ('IGV 18% Cobrado',          resumen['total_igv']),
        ('Costo Envíos',             resumen['gasto_envio']),
        ('Costo Productos',          resumen['costo_productos']),
        ('Ganancia Bruta',           resumen['ganancia_bruta']),
    ]:
        ws2.append([concepto, float(monto)])
        ws2.cell(ws2.max_row, 2).number_format = '"S/ "#,##0.00'

    ws2.append([])
    ws2.append([f'Comprobantes procesados: {resumen["total_comprobantes"]}'])
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre = f'ganancias_consolidado_{fecha_ini.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Exportación Excel — Detallado (por ítem)
# ─────────────────────────────────────────────────────────────────────────────

@reportes_bp.route('/ganancias/exportar-detallado')
@login_required
@requiere_permiso('reportes.exportar')
def exportar_ganancias_detallado():
    """Exporta reporte detallado a Excel (.xlsx) con una fila por ítem de comprobante."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fecha_ini     = _parse_date(request.args.get('fecha_ini', '')) or date.today().replace(day=1)
    fecha_fin     = _parse_date(request.args.get('fecha_fin', '')) or date.today()
    tipo_filtro   = request.args.get('tipo', '')
    fuente_filtro = request.args.get('fuente', '')

    _TIPOS_DETALLADO = _TIPOS_VENTA + ('NOTA_CREDITO',)
    query = (
        Comprobante.query
        .filter(
            Comprobante.tipo_comprobante.in_(_TIPOS_DETALLADO),
            Comprobante.estado.in_(_ESTADOS_VALIDOS),
            func.date(Comprobante.fecha_emision) >= fecha_ini,
            func.date(Comprobante.fecha_emision) <= fecha_fin,
        )
        .order_by(Comprobante.fecha_emision)
    )
    if tipo_filtro in _TIPOS_VENTA:
        query = query.filter(Comprobante.tipo_comprobante == tipo_filtro)
    if fuente_filtro in _FUENTE_FILTROS:
        query = _FUENTE_FILTROS[fuente_filtro](query)

    mapa_costos = _build_mapa_costos()
    comprobantes = query.all()

    _COLOR_HDR = 'FF1e3a5f'
    _COLOR_TOT = 'FFE8F4FD'
    _COLOR_ALT = 'FFF7FAFD'
    BORDER_SIDE = Side(style='thin', color='BDD7EE')
    CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                         top=BORDER_SIDE, bottom=BORDER_SIDE)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Detallado'

    hdrs = [
        'Fuente', 'N° Orden', 'Serie', 'Correlativo', 'Tipo', 'Fecha', 'Cliente',
        'SKU', 'Producto', 'Cantidad', 'Precio Unit. (S/)',
        'Ingreso Item (S/)', 'Costo Unit. USD', 'Costo Unit. (S/)',
        'Costo Total (S/)', 'Ganancia (S/)', 'Margen %',
    ]
    ws.append(hdrs)
    ws.row_dimensions[1].height = 26
    for col_idx in range(1, len(hdrs) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color='FFFFFFFF', size=11)
        cell.fill      = PatternFill('solid', fgColor=_COLOR_HDR)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = CELL_BORDER

    fmt_soles = '"S/ "#,##0.00'
    fmt_usd   = '"$"#,##0.0000'
    fmt_pct   = '0.00"%"'

    total_ingreso = total_costo = total_ganancia = 0.0
    n_filas = 0

    for comp in comprobantes:
        fecha_str    = comp.fecha_emision.strftime('%d/%m/%Y') if comp.fecha_emision else ''
        cliente_nom  = comp.cliente.nombre_completo if comp.cliente else '—'
        tc_comp      = get_tipo_cambio(comp.fecha_emision) or 3.75

        es_nc = comp.tipo_comprobante == 'NOTA_CREDITO'
        envio_item = None
        for item in comp.items:
            sku = (item.producto_sku or '').strip()
            if not sku:
                continue
            if sku == 'ENVIO':
                envio_item = item
                continue

            cantidad = float(item.cantidad or 1)
            if es_nc:
                ingreso_item   = -float(item.subtotal_con_igv or 0)
                costo_unit_usd = 0.0
                costo_unit_pen = 0.0
                costo_total    = 0.0
            else:
                skus = extraer_skus_base(sku)
                costo_unit_usd = sum(mapa_costos.get(s, 0.0) for s in skus)
                costo_unit_pen = round(costo_unit_usd * tc_comp, 2)
                costo_total    = round(costo_unit_pen * cantidad, 2)
                ingreso_item   = float(item.subtotal_con_igv or 0)
            ganancia = round(ingreso_item - costo_total, 2)
            margen   = round(ganancia / ingreso_item * 100, 2) if ingreso_item else 0.0

            _partes = comp.numero_completo.split('-', 1)
            row_idx = ws.max_row + 1
            ws.append([
                _detectar_fuente(comp.numero_orden),
                comp.numero_orden or '—',
                _partes[0] if len(_partes) > 0 else comp.numero_completo,
                _partes[1] if len(_partes) > 1 else '',
                comp.tipo_comprobante,
                fecha_str,
                cliente_nom,
                sku,
                item.producto_nombre,
                cantidad,
                float(item.precio_unitario_con_igv or 0),
                ingreso_item,
                round(costo_unit_usd, 4),
                costo_unit_pen,
                costo_total,
                ganancia,
                margen,
            ])
            ws.row_dimensions[row_idx].height = 17
            is_alt = (n_filas % 2 == 0)
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = CELL_BORDER
                if is_alt:
                    cell.fill = PatternFill('solid', fgColor=_COLOR_ALT)
                # cols: 1=Fuente 2=N°Orden 3=Serie 4=Correlativo 5=Tipo 6=Fecha 7=Cliente
                #       8=SKU 9=Producto 10=Cantidad 11=PrecioUnit 12=IngresoItem
                #       13=CostoUSD 14=CostoPen 15=CostoTotal 16=Ganancia 17=Margen
                if col_idx in (1, 2, 3, 4, 5, 6, 10):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx in (8, 9):
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col_idx == 11:
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.number_format = fmt_soles
                elif col_idx == 12:
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.number_format = fmt_soles
                elif col_idx == 13:
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.number_format = fmt_usd
                elif col_idx in (14, 15, 16):
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.number_format = fmt_soles
                elif col_idx == 17:
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.number_format = fmt_pct

            total_ingreso  += ingreso_item
            total_costo    += costo_total
            total_ganancia += ganancia
            n_filas += 1

        # Fila ENVIO — usa el item ENVIO si existe, sino comp.costo_envio
        envio_monto = 0.0
        if envio_item:
            envio_monto = float(envio_item.subtotal_con_igv or 0)
        elif comp.costo_envio and float(comp.costo_envio) > 0:
            envio_monto = float(comp.costo_envio)
        if envio_monto > 0:
            row_idx = ws.max_row + 1
            _ep = comp.numero_completo.split('-', 1)
            ws.append([
                _detectar_fuente(comp.numero_orden),
                comp.numero_orden or '—',
                _ep[0] if len(_ep) > 0 else comp.numero_completo,
                _ep[1] if len(_ep) > 1 else '',
                comp.tipo_comprobante,
                fecha_str,
                cliente_nom,
                'ENVIO',
                'Costo de Envío',
                1,
                envio_monto,
                envio_monto,
                0.0, 0.0, envio_monto, 0.0, 0.0,
            ])
            ws.row_dimensions[row_idx].height = 17
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = CELL_BORDER
                cell.fill   = PatternFill('solid', fgColor='FFFFF9C4')  # amarillo suave
                if col_idx in (11, 12, 14, 15, 16):
                    cell.number_format = fmt_soles
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx == 17:
                    cell.number_format = fmt_pct
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center' if col_idx in (1,2,3,4,5,6,10) else 'left',
                                               vertical='center')
            total_ingreso += envio_monto
            total_costo   += envio_monto
            n_filas += 1

    # Fila de totales
    total_row = ws.max_row + 1
    ws.row_dimensions[total_row].height = 22
    # cols: 1=Fuente 2=N°Orden 3=Serie 4=Correlativo 5=Tipo 6=Fecha 7=Cliente 8=SKU 9=Producto
    #       10=Cantidad 11=PrecioUnit 12=IngresoItem 13=CostoUSD 14=CostoPen
    #       15=CostoTotal 16=Ganancia 17=Margen
    totales = ['TOTAL', '', '', '', f'{n_filas} ítems', '', '', '', '', '',
               '', round(total_ingreso, 2), '', '', round(total_costo, 2),
               round(total_ganancia, 2),
               round(total_ganancia / total_ingreso * 100, 2) if total_ingreso else 0]
    ws.append(totales)
    for col_idx, cell in enumerate(ws[total_row], start=1):
        cell.font   = Font(bold=True, size=11)
        cell.fill   = PatternFill('solid', fgColor=_COLOR_TOT)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center' if col_idx <= 6 else 'right',
                                   vertical='center')
        if col_idx in (12, 15, 16):
            cell.number_format = fmt_soles
        elif col_idx == 17:
            cell.number_format = fmt_pct

    # Anchos: Fuente, N°Orden, Serie, Correlativo, Tipo, Fecha, Cliente, SKU, Producto, ...
    col_widths = [12, 20, 10, 12, 10, 12, 24, 16, 34, 10, 16, 16, 14, 14, 14, 14, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre = f'ganancias_detallado_{fecha_ini.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _calcular_resumen(comprobantes: list, mapa_costos: dict) -> dict:
    total_ingresos  = Decimal('0')
    total_igv       = Decimal('0')
    base_imponible  = Decimal('0')
    gasto_envio     = Decimal('0')
    costo_productos = Decimal('0')

    for c in comprobantes:
        total_ingresos  += c.total or Decimal('0')
        total_igv       += c.total_igv or Decimal('0')
        base_imponible  += c.total_operaciones_gravadas or Decimal('0')
        gasto_envio     += c.costo_envio or Decimal('0')
        costo_productos += _costo_comprobante(c, mapa_costos)

    ganancia_bruta = total_ingresos - costo_productos - gasto_envio

    return {
        'total_ingresos':    total_ingresos,
        'total_igv':         total_igv,
        'base_imponible':    base_imponible,
        'gasto_envio':       gasto_envio,
        'costo_productos':   costo_productos,
        'ganancia_bruta':    ganancia_bruta,
        'total_comprobantes': len(comprobantes),
        'margen_pct': (
            round(float(ganancia_bruta / total_ingresos * 100), 1)
            if total_ingresos > 0 else 0
        ),
    }


def _costo_comprobante(comp: Comprobante, mapa_costos: dict) -> Decimal:
    tc = get_tipo_cambio(comp.fecha_emision) or 3.75
    total = Decimal('0')
    for item in comp.items:
        sku = (item.producto_sku or '').strip()
        if not sku or sku == 'ENVIO':
            continue
        skus = extraer_skus_base(sku)
        costo_unit_usd = sum(mapa_costos.get(s, 0.0) for s in skus)
        costo_unit_pen = Decimal(str(round(costo_unit_usd * tc, 4)))
        total += costo_unit_pen * (item.cantidad or Decimal('1'))
    return total


def _detectar_fuente(numero_orden: str | None) -> str:
    if not numero_orden:
        return 'Manual'
    n = str(numero_orden).strip()
    if n.startswith('2') and len(n) >= 14:
        return 'MercadoLibre'
    if n.startswith('3') and len(n) >= 10:
        return 'Falabella'
    return 'WooCommerce'


def _enriquecer_fila(comp: Comprobante, mapa_costos: dict) -> dict:
    costo_prods   = _costo_comprobante(comp, mapa_costos)
    total         = comp.total or Decimal('0')
    costo_envio   = comp.costo_envio or Decimal('0')
    descuento     = comp.descuento or Decimal('0')
    ganancia      = total - costo_prods - costo_envio
    margen_pct    = round(float(ganancia / total * 100), 1) if total > 0 else 0

    return {
        'id':              comp.id,
        'numero_completo': comp.numero_completo,
        'numero_orden':    comp.numero_orden or '—',
        'fuente':          _detectar_fuente(comp.numero_orden),
        'tipo_comprobante': comp.tipo_comprobante,
        'fecha_emision':   comp.fecha_emision.strftime('%d/%m/%Y') if comp.fecha_emision else '',
        'cliente_nombre':  comp.cliente.nombre_completo if comp.cliente else '—',
        'estado':          comp.estado,
        'base_imponible':  comp.total_operaciones_gravadas or Decimal('0'),
        'total_igv':       comp.total_igv or Decimal('0'),
        'descuento':       descuento,
        'costo_envio':     costo_envio,
        'total':           total,
        'costo_productos': costo_prods,
        'ganancia_bruta':  ganancia,
        'margen_pct':      margen_pct,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Registro de Envío SUNAT
# ─────────────────────────────────────────────────────────────────────────────

@reportes_bp.route('/envio-sunat')
@login_required
@requiere_permiso('reportes.ver')
def envio_sunat():
    """Formulario de parámetros para el Registro de Envío SUNAT."""
    hoy       = date.today()
    fecha_ini = request.args.get('fecha_ini') or hoy.replace(day=1).strftime('%Y-%m-%d')
    fecha_fin = request.args.get('fecha_fin') or hoy.strftime('%Y-%m-%d')
    return render_template('reportes/envio_sunat.html',
                           fecha_ini=fecha_ini, fecha_fin=fecha_fin)


@reportes_bp.route('/envio-sunat/preview')
@login_required
@requiere_permiso('reportes.ver')
def envio_sunat_preview():
    """Devuelve JSON con los datos del Registro de Envío SUNAT para previsualización."""
    from flask import jsonify
    fecha_ini_str = request.args.get('fecha_ini', '')
    fecha_fin_str = request.args.get('fecha_fin', '')
    if not fecha_ini_str or not fecha_fin_str:
        return jsonify({'error': 'Debe indicar rango de fechas.'}), 400

    comprobantes = (Comprobante.query
                    .filter(Comprobante.estado != 'BORRADOR',
                            Comprobante.fecha_emision.between(
                                f'{fecha_ini_str} 00:00:00',
                                f'{fecha_fin_str} 23:59:59'))
                    .order_by(Comprobante.fecha_emision)
                    .all())

    TIPO_NOMBRE = {
        'BOLETA':       'Boleta de Venta',
        'NOTA_CREDITO': 'Nota de Crédito',
        'FACTURA':      'Factura de Venta',
    }
    ESTADO_LETRA = {
        'ACEPTADO': 'A', 'ENVIADO': 'E', 'RECHAZADO': 'R',
        'PENDIENTE': 'P', 'BORRADOR': 'B',
    }

    rows = []
    for comp in comprobantes:
        ref_num = ''
        if comp.tipo_comprobante in ('NOTA_CREDITO', 'NOTA_DEBITO') and comp.comprobante_ref:
            ref_num = comp.comprobante_ref.numero_completo
        rows.append({
            'tipo':      TIPO_NOMBRE.get(comp.tipo_comprobante, comp.tipo_comprobante),
            'numero':    comp.numero_completo,
            'fecha':     comp.fecha_emision.strftime('%d/%m/%Y') if comp.fecha_emision else '',
            'estado':    ESTADO_LETRA.get(comp.estado, comp.estado[:1] if comp.estado else ''),
            'doc_ref':   ref_num,
            'respuesta': comp.mensaje_sunat or '',
            'cod_res':   str(comp.codigo_sunat) if comp.codigo_sunat is not None else '',
            'enviado':   1 if comp.estado in ('ENVIADO', 'ACEPTADO') else 0,
        })

    return jsonify({'total': len(rows), 'rows': rows})


@reportes_bp.route('/envio-sunat/exportar')
@login_required
@requiere_permiso('reportes.ver')
def envio_sunat_exportar():
    """Genera y descarga el Excel de Registro de Envío SUNAT."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    fecha_ini_str = request.args.get('fecha_ini', '')
    fecha_fin_str = request.args.get('fecha_fin', '')

    if not fecha_ini_str or not fecha_fin_str:
        from flask import flash, redirect, url_for
        flash('Debe indicar rango de fechas.', 'danger')
        return redirect(url_for('reportes.envio_sunat'))

    comprobantes = (Comprobante.query
                    .filter(Comprobante.estado != 'BORRADOR',
                            Comprobante.fecha_emision.between(
                                f'{fecha_ini_str} 00:00:00',
                                f'{fecha_fin_str} 23:59:59'))
                    .order_by(Comprobante.fecha_emision)
                    .all())

    TIPO_NOMBRE = {
        'BOLETA':       'BOLETA DE VENTA ELECTRONICA',
        'NOTA_CREDITO': 'NOTA CREDITO ELECTRONICA',
        'FACTURA':      'FACTURA DE VENTA ELECTRONICA',
    }
    ESTADO_LETRA = {
        'ACEPTADO': 'A', 'ENVIADO': 'E', 'RECHAZADO': 'R',
        'PENDIENTE': 'P', 'BORRADOR': 'B',
    }

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Envío SUNAT'

    empresa_ruc   = current_app.config.get('EMPRESA_RUC', '')
    empresa_razon = current_app.config.get('EMPRESA_RAZON_SOCIAL', '')

    thin        = Side(style='thin')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill('solid', fgColor='1F4E79')
    fill_alt    = PatternFill('solid', fgColor='EEF3FB')
    align_c = Alignment(horizontal='center', vertical='center')
    align_l = Alignment(horizontal='left',   vertical='center')

    ws['A1'] = date.today().strftime('%d/%m/%Y')
    ws['A2'] = f'RUC: {empresa_ruc}'
    ws['A3'] = f'Razón Social: {empresa_razon}'
    ws['A4'] = f'Período: {fecha_ini_str} al {fecha_fin_str}'
    ws['A1'].font = Font(name='Arial', bold=True, size=10)
    for r in range(2, 5):
        ws.cell(r, 1).font = Font(name='Arial', size=9)

    HEADERS = ['N.', 'TTienda', 'Documento', 'Documento',
               'Fecha', 'Estado', 'NumeroOriginal', 'FechaOriginal',
               'Documento Ref', 'Respuesta Sunat', 'CodRes', 'Enviado']
    ROW_HDR = 6
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=ROW_HDR, column=ci, value=h)
        cell.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill      = fill_header
        cell.alignment = align_c
        cell.border    = border
    ws.row_dimensions[ROW_HDR].height = 18

    col_widths = [6, 10, 28, 18, 12, 8, 14, 12, 18, 60, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for n, comp in enumerate(comprobantes, 1):
        tipo_nombre  = TIPO_NOMBRE.get(comp.tipo_comprobante, comp.tipo_comprobante)
        estado_letra = ESTADO_LETRA.get(comp.estado, comp.estado[:1] if comp.estado else '')
        ref_num      = ''
        if comp.tipo_comprobante in ('NOTA_CREDITO', 'NOTA_DEBITO') and comp.comprobante_ref:
            ref_num = comp.comprobante_ref.numero_completo
        enviado   = 1 if comp.estado in ('ENVIADO', 'ACEPTADO') else 0
        fecha_str = comp.fecha_emision.strftime('%d/%m/%Y') if comp.fecha_emision else ''

        row_idx = ROW_HDR + n
        ws.append([
            0, 'CENTRAL', tipo_nombre, comp.numero_completo, fecha_str,
            estado_letra, '', '',
            ref_num,
            comp.mensaje_sunat or '',
            comp.codigo_sunat or '',
            enviado,
        ])
        ws.row_dimensions[row_idx].height = 14
        is_alt = (n % 2 == 0)
        for ci, cell in enumerate(ws[row_idx], start=1):
            cell.font   = Font(name='Arial', size=9)
            cell.border = border
            if is_alt:
                cell.fill = fill_alt
            cell.alignment = align_c if ci in (1, 2, 5, 6, 7, 8, 11, 12) else align_l

    ws.freeze_panes = ws.cell(row=ROW_HDR + 1, column=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'EnvioSUNAT_{fecha_ini_str}_al_{fecha_fin_str}.xlsx',
    )


def _parse_date(s: str):
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Registro de Ventas e Ingresos — Formato 14.1
# ─────────────────────────────────────────────────────────────────────────────

@reportes_bp.route('/registro-ventas')
@login_required
@requiere_permiso('reportes.ver')
def registro_ventas():
    """Formulario de parámetros para el Registro de Ventas e Ingresos."""
    hoy       = date.today()
    fecha_ini = request.args.get('fecha_ini') or hoy.replace(day=1).strftime('%Y-%m-%d')
    fecha_fin = request.args.get('fecha_fin') or hoy.strftime('%Y-%m-%d')
    return render_template('reportes/registro_ventas.html',
                           fecha_ini=fecha_ini, fecha_fin=fecha_fin)


@reportes_bp.route('/registro-ventas/preview')
@login_required
@requiere_permiso('reportes.ver')
def registro_ventas_preview():
    """Devuelve JSON con los datos del Registro de Ventas para previsualización."""
    from flask import jsonify
    fecha_ini_str = request.args.get('fecha_ini', '')
    fecha_fin_str = request.args.get('fecha_fin', '')
    if not fecha_ini_str or not fecha_fin_str:
        return jsonify({'error': 'Debe indicar rango de fechas.'}), 400

    comprobantes = (Comprobante.query
                    .filter(Comprobante.estado.notin_(['BORRADOR', 'RECHAZADO']),
                            Comprobante.fecha_emision.between(
                                f'{fecha_ini_str} 00:00:00',
                                f'{fecha_fin_str} 23:59:59'))
                    .order_by(Comprobante.tipo_comprobante, Comprobante.fecha_emision)
                    .all())

    def tipo_doc_sunat(comp):
        if comp.tipo_comprobante == 'NOTA_CREDITO':
            return '07'
        if comp.serie and comp.serie.upper().startswith('F'):
            return '01'
        return '03'

    rows = []
    nro = 0
    total_base = total_igv = total_tot = 0.0
    for comp in comprobantes:
        nro += 1
        total_v = float(comp.total or 0)
        if comp.tipo_comprobante == 'NOTA_CREDITO':
            total_v = -abs(total_v)
        b = round(total_v / 1.18, 2)
        i = round(total_v * 18 / 118, 2)
        tc = get_tipo_cambio(comp.fecha_emision) or 0.0
        ref_num = ''
        if comp.tipo_comprobante == 'NOTA_CREDITO' and comp.comprobante_ref:
            ref_num = comp.comprobante_ref.numero_completo
        total_base += b
        total_igv  += i
        total_tot  += total_v
        rows.append({
            'nro':      nro,
            'fecha':    comp.fecha_emision.strftime('%d/%m/%Y') if comp.fecha_emision else '',
            'tipo_doc': tipo_doc_sunat(comp),
            'serie':    comp.serie or '',
            'numero':   str(comp.correlativo).zfill(8),
            'nombre':   comp.cliente.nombre_completo if comp.cliente else '—',
            'base_imp': b,
            'igv':      i,
            'total':    total_v,
            'tc':       round(float(tc), 4) if tc else '',
            'ref_num':  ref_num,
        })

    return jsonify({
        'total':      len(rows),
        'rows':       rows,
        'total_base': round(total_base, 2),
        'total_igv':  round(total_igv, 2),
        'total_tot':  round(total_tot, 2),
    })


@reportes_bp.route('/registro-ventas/exportar')
@login_required
@requiere_permiso('reportes.ver')
def registro_ventas_exportar():
    """Genera y descarga el Excel Formato 14.1."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    fecha_ini_str = request.args.get('fecha_ini', '')
    fecha_fin_str = request.args.get('fecha_fin', '')

    if not fecha_ini_str or not fecha_fin_str:
        from flask import flash, redirect, url_for
        flash('Debe indicar rango de fechas.', 'danger')
        return redirect(url_for('reportes.registro_ventas'))

    comprobantes = (Comprobante.query
                    .filter(Comprobante.estado.notin_(['BORRADOR', 'RECHAZADO']),
                            Comprobante.fecha_emision.between(
                                f'{fecha_ini_str} 00:00:00',
                                f'{fecha_fin_str} 23:59:59'))
                    .order_by(Comprobante.tipo_comprobante, Comprobante.fecha_emision)
                    .all())

    def tipo_doc_sunat(comp):
        if comp.tipo_comprobante == 'NOTA_CREDITO':
            return '07'
        if comp.serie and comp.serie.upper().startswith('F'):
            return '01'
        return '03'

    ID_DOC_MAP = {'DNI': '1', 'RUC': '6', 'CE': '4', 'PAS': '7'}

    def base_imp(total):
        return round(float(total) / 1.18, 2)

    def igv_val(total):
        return round(float(total) * 18 / 118, 2)

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Registro de Ventas'

    empresa_ruc   = current_app.config.get('EMPRESA_RUC', '')
    empresa_razon = current_app.config.get('EMPRESA_RAZON_SOCIAL', '')

    thin        = Side(style='thin')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill('solid', fgColor='1F4E79')
    fill_subtot = PatternFill('solid', fgColor='D9E1F2')
    fill_total  = PatternFill('solid', fgColor='BDD7EE')
    align_c     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_l     = Alignment(horizontal='left',   vertical='center')
    align_r     = Alignment(horizontal='right',  vertical='center')
    fmt_date    = 'DD/MM/YY'
    fmt_num     = '#,##0.00'
    fmt_num4    = '#,##0.0000'

    # Encabezado
    ws.merge_cells('A1:V1')
    ws['A1'] = 'Formato 14.1: Registro de Ventas e Ingresos'
    ws['A1'].font = Font(name='Arial', bold=True, size=12)
    ws['A1'].alignment = align_l

    ws.merge_cells('A2:V2')
    ws['A2'] = f'Rango de Fechas: {fecha_ini_str} a {fecha_fin_str}'
    ws['A2'].font = Font(name='Arial', size=9)

    ws['A3'] = 'RUC:';          ws['B3'] = empresa_ruc
    ws['A4'] = 'Razón Social:'; ws['B4'] = empresa_razon
    for cell in [ws['A3'], ws['A4']]:
        cell.font = Font(name='Arial', bold=True, size=9)
    for cell in [ws['B3'], ws['B4']]:
        cell.font = Font(name='Arial', size=9)

    HEADERS = [
        'Nro.\nCorrelativo', 'Fecha\nEmisión', 'Fecha\nVencim.',
        'Tipo\nDoc.', 'Serie', 'Número',
        'T.Doc.\nIdentidad', 'N° Documento\nIdentidad', 'Nombre / Razón Social',
        'Val. Fact.\nExportación', 'Base Imponible\nOp. Gravada',
        'Exon.', 'Inafecto', 'ISC', 'IGV', 'Otros\nTributos',
        'Importe\nTotal', 'Tipo\nCambio',
        'Ref: Fecha', 'Ref: Tipo', 'Ref: Serie', 'Ref: N° Comprobante',
    ]
    ROW_HDR = 6
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=ROW_HDR, column=ci, value=h)
        cell.font      = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        cell.fill      = fill_header
        cell.alignment = align_c
        cell.border    = border_thin
    ws.row_dimensions[ROW_HDR].height = 30

    col_widths = [6, 10, 10, 6, 7, 11, 8, 14, 35, 10, 12, 7, 8, 7, 10, 8, 10, 8, 10, 6, 8, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Datos ────────────────────────────────────────────────────────────────
    row   = ROW_HDR + 1
    nro   = 0
    ESTABLECIMIENTO = 'CENTRAL'
    grupos_orden    = ['01', '03', '07']
    grupos: dict[str, list] = {}
    for c in comprobantes:
        grupos.setdefault(tipo_doc_sunat(c), []).append(c)

    est_base = est_igv = est_tot = 0.0

    ws.cell(row=row, column=1, value=ESTABLECIMIENTO).font = Font(name='Arial', bold=True, size=8)
    ws.row_dimensions[row].height = 14
    row += 1

    for td in grupos_orden:
        if td not in grupos:
            continue
        grp_base = grp_igv = grp_tot = 0.0

        for comp in grupos[td]:
            nro += 1
            tc      = get_tipo_cambio(comp.fecha_emision) or 0.0
            total_v = float(comp.total or 0)
            if comp.tipo_comprobante == 'NOTA_CREDITO':
                total_v = -abs(total_v)
            b = base_imp(total_v)
            i = igv_val(total_v)
            grp_base += b; grp_igv += i; grp_tot += total_v

            ref_fecha = ref_tipo = ref_serie = ref_num = ''
            if comp.tipo_comprobante == 'NOTA_CREDITO' and comp.comprobante_ref:
                ref = comp.comprobante_ref
                ref_fecha = ref.fecha_emision.date() if ref.fecha_emision else ''
                ref_tipo  = tipo_doc_sunat(ref)
                ref_serie = ref.serie or ''
                ref_num   = str(ref.correlativo).zfill(8)

            cli     = comp.cliente
            id_tipo = ID_DOC_MAP.get(cli.tipo_documento, cli.tipo_documento) if cli else ''
            id_num  = cli.numero_documento if cli else ''
            nombre  = cli.nombre_completo  if cli else ''

            valores = [
                nro,
                comp.fecha_emision.date() if comp.fecha_emision else '',
                comp.fecha_emision.date() if comp.fecha_emision else '',
                tipo_doc_sunat(comp), comp.serie,
                str(comp.correlativo).zfill(8),
                id_tipo, id_num, nombre, '',
                b, '', '', '', i, '', total_v,
                tc if tc else '',
                ref_fecha, ref_tipo, ref_serie, ref_num,
            ]
            for ci, val in enumerate(valores, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font   = Font(name='Arial', size=8)
                cell.border = border_thin
                cell.alignment = align_l
                if ci in (2, 3, 19):
                    cell.number_format = fmt_date; cell.alignment = align_c
                elif ci in (11, 15, 17):
                    cell.number_format = fmt_num;  cell.alignment = align_r
                elif ci == 18:
                    cell.number_format = fmt_num4; cell.alignment = align_r
                elif ci in (4, 5, 6, 7, 20, 21, 22):
                    cell.alignment = align_c
            ws.row_dimensions[row].height = 13
            row += 1

        # Subtotal por tipo doc
        ws.cell(row=row, column=9, value=f'Totales por {td} :').font = Font(name='Arial', bold=True, size=8)
        for ci, val in [(11, grp_base), (15, grp_igv), (17, grp_tot)]:
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = Font(name='Arial', bold=True, size=8)
            cell.fill = fill_subtot; cell.border = border_thin
            cell.number_format = fmt_num; cell.alignment = align_r
        ws.row_dimensions[row].height = 14
        row += 1
        est_base += grp_base; est_igv += grp_igv; est_tot += grp_tot

    # Subtotal establecimiento
    ws.cell(row=row, column=9, value=f'Totales por {ESTABLECIMIENTO}:').font = Font(name='Arial', bold=True, size=8)
    for ci, val in [(11, est_base), (15, est_igv), (17, est_tot)]:
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = Font(name='Arial', bold=True, size=8)
        cell.fill = fill_total; cell.border = border_thin
        cell.number_format = fmt_num; cell.alignment = align_r
    ws.row_dimensions[row].height = 14
    row += 1

    # Gran total
    ws.cell(row=row, column=9, value='Totales:').font = Font(name='Arial', bold=True, size=9)
    for ci, val in [(11, est_base), (15, est_igv), (17, est_tot)]:
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = fill_total; cell.border = border_thin
        cell.number_format = fmt_num; cell.alignment = align_r
    ws.row_dimensions[row].height = 14

    ws.freeze_panes = ws.cell(row=ROW_HDR + 1, column=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'RegistroVentas_{fecha_ini_str}_al_{fecha_fin_str}.xlsx'
    )
